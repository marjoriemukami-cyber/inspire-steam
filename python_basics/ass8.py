#register new students; id,name,course,phone
#display students
# assign new course,grade
#store in ex file

from openpyxl import Workbook, load_workbook
import os

FILE_NAME = "students.xlsx"

# Create file if it doesn't exist
def create_file():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append(["ID", "Name", "Course", "Phone", "Grade"])
        wb.save(FILE_NAME)

# Register new student
def register_student():
    id = input("Enter Student ID: ")
    name = input("Enter Name: ")
    course = input("Enter Course: ")
    phone = input("Enter Phone: ")

    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append([id, name, course, phone, ""])
    wb.save(FILE_NAME)
    print("✅ Student registered!\n")

# Display students
def display_students():
    wb = load_workbook(FILE_NAME)
    ws = wb.active

    print("\n📋 Student List")
    for row in ws.iter_rows(values_only=True):
        print(row)
    print()

# Assign new course & grade
def assign_course_grade():
    student_id = input("Enter Student ID: ")
    new_course = input("Enter New Course: ")
    grade = input("Enter Grade: ")

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    for row in ws.iter_rows():
        if row[0].value == student_id:
            row[2].value = new_course
            row[4].value = grade
            wb.save(FILE_NAME)
            print("✅ Updated successfully!\n")
            return

    print("❌ Student not found!\n")

# Menu
def main():
    create_file()

    while True:
        print("1. Register Student")
        print("2. Display Students")
        print("3. Assign Course & Grade")
        print("4. Exit")

        choice = input("Choose option: ")

        if choice == "1":
            register_student()
        elif choice == "2":
            display_students()
        elif choice == "3":
            assign_course_grade()
        elif choice == "4":
            break
        else:
            print("Invalid choice!\n")

main()